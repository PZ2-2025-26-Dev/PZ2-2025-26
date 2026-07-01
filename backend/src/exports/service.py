from io import BytesIO
from uuid import UUID

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from src.items.models import Category, Item
from src.items.schemas import ItemSearch
from src.items.service import ItemService
from src.loans.constants import ReturnCondition
from src.loans.models import Loan
from src.locations.models import Location
from src.users.models import User


class ExportService:
    def __init__(self, db):
        self.db = db

    def get_items(
        self,
        search: str | None = None,
        status: str | None = None,
        category: str | None = None,
    ):
        stmt = select(Item).order_by(Item.id)

        if status:
            stmt = stmt.where(Item.status == status)

        if category:
            stmt = stmt.where(Item.category.has(name=category))

        if search:
            like = f"%{search}%"
            stmt = stmt.where((Item.name.ilike(like)) | (Item.description.ilike(like)))

        return self.db.execute(stmt).scalars().all()

    def export_items_xlsx(self, data: ItemSearch):
        stmt = ItemService(self.db)._build_items_query(data)
        stmt = stmt.options(
            selectinload(Item.category),
            selectinload(Item.location),
            selectinload(Item.owner),
        )

        sort_field_map = {
            "id": Item.id,
            "name": Item.name,
            "status": Item.status,
            "category": Category.name,
            "location": Location.name,
            "owner": User.last_name,
        }

        # Parse multiple sort criteria: "name:asc,status:desc,category:asc"
        sort_criterias = []
        if data.sort:
            parts = data.sort.split(",")
            for part in parts:
                if ":" in part:
                    field, order = part.split(":", 1)
                else:
                    field, order = part, "asc"

                field = field.strip().lower()
                order = order.strip().lower()
                if field in sort_field_map:
                    sort_criterias.append((field, order))

        joined_tables = set()
        for field, _ in sort_criterias:
            if field == "category" and "category" not in joined_tables:
                stmt = stmt.join(Item.category, isouter=True)
                joined_tables.add("category")
            elif field == "location" and "location" not in joined_tables:
                stmt = stmt.join(Item.location, isouter=True)
                joined_tables.add("location")
            elif field == "owner" and "owner" not in joined_tables:
                stmt = stmt.join(Item.owner, isouter=True)
                joined_tables.add("owner")

        order_by_clauses = []
        for field, order in sort_criterias:
            column = sort_field_map.get(field, Item.id)
            if order == "desc":
                order_by_clauses.append(column.desc())
            else:
                order_by_clauses.append(column.asc())

        if not order_by_clauses:
            order_by_clauses.append(Item.name.asc())

        stmt = stmt.order_by(*order_by_clauses)

        items = self.db.execute(stmt).scalars().all()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Items"

        worksheet.append(
            [
                "ID",
                "Name",
                "Inventory Number",
                "Category",
                "Location",
                "Owner",
                "Status",
                "Description",
            ]
        )

        for item in items:
            worksheet.append(
                [
                    item.id,
                    item.name,
                    item.oldID or "",
                    item.category.name if item.category else "",
                    item.location.name if item.location else "",
                    f"{item.owner.first_name} {item.owner.last_name}".strip() if item.owner else "",
                    item.status.value,
                    item.description or "",
                ]
            )

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="items.xlsx"',
            },
        )

    def export_item_report_xlsx(self, item_uuid: UUID):
        item_stmt = (
            select(Item)
            .where(Item.id == item_uuid)
            .options(
                selectinload(Item.category),
                selectinload(Item.location),
                selectinload(Item.owner),
            )
        )
        item = self.db.execute(item_stmt).scalar_one_or_none()
        if not item:
            raise ValueError("Item not found")

        stats_stmt = select(
            func.count(Loan.id).label("total_loans"),
            func.count(func.distinct(func.coalesce(Loan.user_id, Loan.guest_id))).label("unique_borrowers"),
            func.avg(func.datediff(Loan.returned_at, Loan.borrowed_at)).label("avg_duration"),
            func.sum(func.if_(Loan.return_condition == ReturnCondition.BROKEN, 1, 0)).label("broken_count"),
        ).where(Loan.item_id == item.id)
        stats = self.db.execute(stats_stmt).one()

        total_loans = stats.total_loans or 0
        unique_borrowers = stats.unique_borrowers or 0
        broken_count = stats.broken_count or 0

        avg_duration_str = "N/A"
        if stats.avg_duration is not None:
            days = int(stats.avg_duration)
            avg_duration_str = f"{days} dni" if days > 0 else "mniej niż 1 dzień"
        else:
            avg_duration_str = "N/A"

        failure_rate = f"{(broken_count / total_loans) * 100:.1f}%" if total_loans > 0 else "0.0%"

        history_stmt = select(Loan).where(Loan.item_id == item.id).order_by(Loan.created_at.desc())
        loans_history = self.db.execute(history_stmt).scalars().all()
        print(loans_history)

        # # 4. Budowanie pliku Excel
        workbook = Workbook()

        ws_summary = workbook.active
        ws_summary.title = "Summary"

        bold_font = Font(bold=True)
        header_font = Font(bold=True, size=14)

        # Sekcja: Dane przedmiotu
        ws_summary.append(["ITEM REPORT"])
        ws_summary.cell(1, 1).font = header_font
        ws_summary.append([])

        ws_summary.append(["Item ID:", item.id])
        ws_summary.append(["Name:", item.name])
        ws_summary.append(["Inventory Number:", item.oldID or ""])
        ws_summary.append(["Category:", item.category.name if item.category else ""])
        ws_summary.append(["Current Status:", item.status.value])

        for row in range(3, 8):
            ws_summary.cell(row, 1).font = bold_font

        ws_summary.append([])
        ws_summary.append(["METRICS"])
        ws_summary.cell(9, 1).font = header_font
        ws_summary.append([])

        # Sekcja: Metryki
        ws_summary.append(["Total Loans:", total_loans])
        ws_summary.append(["Unique Borrowers:", unique_borrowers])
        ws_summary.append(["Average Loan Duration:", avg_duration_str])
        ws_summary.append(["Damaged Returns Count:", broken_count])
        ws_summary.append(["Failure Rate:", failure_rate])

        for row in range(11, 16):
            ws_summary.cell(row, 1).font = bold_font

        # --- ZAKŁADKA 2: Historia wypożyczeń ---
        ws_history = workbook.create_sheet(title="Loan History")
        ws_history.append(
            [
                "Loan ID",
                "Status",
                "Created At",
                "Borrowed At",
                "Returned At",
                "Borrower Type",
                "Borrower ID",
                "Return Condition",
                "Return Note",
            ]
        )
        ws_history.row_dimensions[1].font = bold_font

        for loan in loans_history:
            borrower_type = "User" if loan.user_id else "Guest"
            borrower_id = loan.user_id if loan.user_id else loan.guest_id

            ws_history.append(
                [
                    loan.id,
                    loan.status.value,
                    loan.created_at.strftime("%Y-%m-%d %H:%M") if loan.created_at else "",
                    loan.borrowed_at.strftime("%Y-%m-%d %H:%M") if loan.borrowed_at else "",
                    loan.returned_at.strftime("%Y-%m-%d %H:%M") if loan.returned_at else "",
                    borrower_type,
                    borrower_id,
                    loan.return_condition.value if loan.return_condition else "",
                    loan.return_note or "",
                ]
            )

        # Zapisuwanie strumienia i zwrot pliku
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        filename = f"report_item_{item_uuid}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )
